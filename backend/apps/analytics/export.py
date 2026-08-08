"""Экспорт отчётов в PDF и Excel (SPEC §9.5, §12.3, issue #31).

Генерация агрегированных отчётов с фирменным стилем (BRANDBOOK §10.1).
Права: HR и руководитель. Каждая выгрузка фиксируется в audit-log.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

from apps.users.models import Role

from .services import build_company_dashboard

if TYPE_CHECKING:
    from apps.users.models import User


class ExportNotAllowed(Exception):
    """Недостаточно прав для экспорта (SPEC §2.2)."""


def _check_export_permission(actor: User) -> None:
    """Проверить право на экспорт (HR/руководитель, SPEC §2.2)."""
    if not actor.has_any_role(Role.Code.HR.value, Role.Code.MANAGER.value):
        raise ExportNotAllowed("Экспорт доступен только HR и руководителям")


def _audit_export(actor: User, fmt: str) -> None:
    """Зафиксировать выгрузку в audit-log (SPEC §12.3)."""
    from apps.audit.services import log_action

    log_action(
        actor=actor,
        action="export.report",
        target_type="analytics.dashboard",
        target_id="company",
        details={"format": fmt},
    )


# Фирменные цвета из BRANDBOOK §3.1 (для PDF — числовые RGB 0-1).
_PRIMARY_RGB = (0.231, 0.357, 0.863)  # #3B5BDB


def export_company_dashboard_pdf(*, actor: User) -> bytes:
    """Сгенерировать PDF-отчёт дашборда компании (SPEC §9.5).

    Возвращает байты PDF. Требует права HR/руководитель.
    """
    from reportlab.lib.pagesizes import A4  # type: ignore[import-untyped]
    from reportlab.lib.units import mm  # type: ignore[import-untyped]
    from reportlab.pdfgen import canvas as rl_canvas  # type: ignore[import-untyped]

    _check_export_permission(actor)
    data = build_company_dashboard()
    _audit_export(actor, "pdf")

    buffer = io.BytesIO()
    c = rl_canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Заголовок (фирменный стиль).
    c.setFillColorRGB(*_PRIMARY_RGB)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(20 * mm, height - 25 * mm, "Vektor — отчёт по компании")

    c.setFillColorRGB(0.1, 0.11, 0.14)  # text-primary
    c.setFont("Helvetica", 11)
    y = height - 45 * mm
    labels = [
        f"Всего сотрудников: {data['total_employees']}",
        f"С завершённой оценкой: {data['assessed_employees']}",
        f"Охват оценки: {data['assessment_coverage']}%",
        f"Средний балл: {data['average_score']}",
        f"Всего циклов: {data['total_cycles']}",
    ]
    for line in labels:
        c.drawString(20 * mm, y, line)
        y -= 7 * mm

    c.save()
    return buffer.getvalue()


def export_company_dashboard_xlsx(*, actor: User) -> bytes:
    """Сгенерировать Excel-отчёт дашборда компании (SPEC §9.5).

    Возвращает байты XLSX. Требует права HR/руководитель.
    """
    from openpyxl import Workbook

    _check_export_permission(actor)
    data = build_company_dashboard()
    _audit_export(actor, "xlsx")

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # wb.active всегда возвращает активный лист
    ws.title = "Отчёт по компании"

    # Заголовки и данные.
    ws.append(["Метрика", "Значение"])
    ws.append(["Всего сотрудников", data["total_employees"]])
    ws.append(["С завершённой оценкой", data["assessed_employees"]])
    ws.append(["Охват оценки, %", data["assessment_coverage"]])
    ws.append(["Средний балл", data["average_score"]])
    ws.append(["Всего циклов", data["total_cycles"]])

    # Стиль заголовка.
    from openpyxl.styles import Font

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
